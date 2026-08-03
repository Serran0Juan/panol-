"""Capa de acceso a datos, mapeada contra el Google Sheet real del pañol.

Pestañas que usa (nombres tal cual están en la planilla):
  - "Inventario"    : catálogo de productos (fuente de verdad del stock)
  - "Vales APP"     : cabecera de cada movimiento (un vale = una entrega/préstamo)
  - "Registro APP"  : renglones de cada vale (un renglón por producto)
  - "Parametros"    : listas desplegables (tipos de movimiento, sectores, unidades...)
  - "Plano Pañol"   : catálogo de las 35 estanterías y qué guarda cada una
  - "Usuarios"      : quién puede entrar a la app (la crea la app si no existe)
  - "Reclamos"      : pedidos/quejas de los operarios (la crea la app si no existe)

Si no hay credenciales configuradas todavía, trabaja contra una copia local de la
planilla (_devdata/) para poder probar sin tocar los datos reales.
"""

import datetime as dt
import os
import re
import time
from pathlib import Path

import pandas as pd
import streamlit as st

DEVDATA_DIR = Path(__file__).parent / "_devdata"
SEED_XLSX = DEVDATA_DIR / "seed_sheet.xlsx"

# ── Hora local ──────────────────────────────────────────────────────────────
# El servidor de Streamlit Cloud corre en UTC, así que dt.datetime.now() daba
# la hora de Londres: un vale cargado a las 9 de la mañana quedaba registrado a
# las 12. Toda fecha que la app escribe o compara pasa por acá.
# Argentina está fija en UTC-3 desde 2009, así que si el sistema no trae la base
# de husos horarios alcanza con el desfasaje fijo.
try:
    from zoneinfo import ZoneInfo

    ZONA = ZoneInfo("America/Argentina/Buenos_Aires")
except Exception:  # noqa: BLE001  (falta tzdata: caemos al offset fijo)
    ZONA = dt.timezone(dt.timedelta(hours=-3))

FORMATO_FECHA_HORA = "%Y-%m-%d %H:%M:%S"


def ahora() -> dt.datetime:
    """Fecha y hora de La Plata, sin la zona pegada.

    Se devuelve "ingenua" a propósito: en la planilla las fechas están guardadas
    sin zona, y mezclar fechas con y sin zona no se puede comparar.
    """
    return dt.datetime.now(ZONA).replace(tzinfo=None)


def ahora_texto() -> str:
    """La hora local lista para escribir en la planilla."""
    return ahora().strftime(FORMATO_FECHA_HORA)


def hoy() -> dt.date:
    return ahora().date()


# ── Aguantar los tropiezos de la API de Google ──────────────────────────────
# Google limita las consultas por minuto. Registrar varias devoluciones seguidas
# dispara muchas lecturas juntas —cada escritura vacía la caché y la pantalla
# siguiente vuelve a leer todas las hojas— y se llega a ese tope. El pedido
# vuelve con un 429 que se arregla solo esperando unos segundos.
REINTENTOS = 4
ESPERA_BASE = 1.5  # segundos; después 3, 6...

CODIGOS_PASAJEROS = (429, 500, 502, 503, 504)


class LimiteDeGoogle(RuntimeError):
    """Se superó el tope de consultas por minuto de Google Sheets."""


def _con_reintentos(operacion):
    """Corre algo contra la planilla, reintentando si el error es pasajero."""
    from gspread.exceptions import APIError

    for intento in range(REINTENTOS):
        try:
            return operacion()
        except APIError as e:
            codigo = getattr(e, "code", None)
            ultimo = intento == REINTENTOS - 1
            if codigo not in CODIGOS_PASAJEROS:
                raise
            if ultimo:
                if codigo == 429:
                    raise LimiteDeGoogle(
                        "Google está recibiendo demasiadas consultas de la app y "
                        "cortó por un rato. Esperá un minuto y volvé a intentar; "
                        "no se perdió nada de lo que ya habías registrado."
                    ) from e
                raise
            time.sleep(ESPERA_BASE * (2 ** intento))


HOJA_INVENTARIO = "Inventario"
HOJA_VALES = "Vales APP"
HOJA_REGISTRO = "Registro APP"
HOJA_PARAMETROS = "Parametros"
HOJA_PLANO = "Plano Pañol"
HOJA_USUARIOS = "Usuarios"
HOJA_RECLAMOS = "Reclamos"
HOJA_ORDENES = "Ordenes"
HOJA_OT_ESTADOS = "OT_Estados"

# nombre interno -> encabezado exacto en la planilla
COLS_INVENTARIO = {
    "id": "Nro/SKU",
    "descripcion": "Descripción del Producto",
    "stock_inicial": "Stock Inicial",
    "stock_actual": "Stock Actual",
    "unidad": "Unidad",
    "ubicacion": "Ubicación",
    "estado_hoja": "Estado/Alerta",
    "stock_minimo": "Stock Minimo",
    "consumo": "Consumo",
    "prestamos_pendiente": "Prestamos Pendiente",
    "precio_unitario": "Precio Unitario",
    "total": "Total",
    "categoria": "Categoria/Area",
    "subcategoria": "Subcategoria",
    "fuente_requerimiento": "Fuente del requerimiento",
}
# Columnas de "Inventario" que la app puede escribir.
# OJO: "stock_actual" NO está y no debe estar. En la planilla es una fórmula
# (=Stock Inicial − Consumo − Préstamos pendientes) que se calcula sola a partir
# de "Registro APP". Si la app escribiera ahí, pisaría la fórmula y rompería el
# cálculo automático de ese producto.
INVENTARIO_ESCRIBIBLES = {"descripcion", "stock_inicial", "unidad", "ubicacion", "stock_minimo",
                          "precio_unitario", "categoria", "subcategoria", "fuente_requerimiento"}

COLS_VALES = ["ID VALE", "FECHA HORA", "TIPO MOVIMIENTO", "SECTOR", "ÁREA / SALA",
              "Receptor / Para Quien", "OBSERVACIONES", "ESTADO VALE", "DIAS RETRASO",
              "REGISTRADO_POR"]
COLS_REGISTRO = ["ID_REGISTRO", "ID_VALE_REF", "ID_ITEM", "CANT", "UNIDAD", "TIPO_MOV",
                 "DESCRIPCIÓN_ITEM", "FECHA_VALE", "OBSERVACIONES", "ESTADO_VALE (auto)",
                 "CANT_DEVUELTA", "ESTADO_RENGLON"]
COLS_USUARIOS = ["EMAIL", "NOMBRE", "ROL", "SECTOR", "ACTIVO", "PASSWORD_HASH"]
COLS_RECLAMOS = ["ID", "FECHA_HORA", "TIPO", "PRODUCTO", "DETALLE", "EMAIL", "NOMBRE", "ESTADO", "RESPUESTA"]

# ── Mantenimiento correctivo: órdenes de trabajo ──
COLS_ORDENES = ["ID_OT", "FECHA_ALTA", "SOLICITANTE", "SOLICITANTE_EMAIL", "AREA",
                "DESCRIPCION", "PRIORIDAD", "SECTOR_ASIGNADO", "ASIGNADO_A", "ESTADO",
                "FECHA_ASIGNACION", "FECHA_CIERRE", "TRABAJO_REALIZADO", "CAUSA",
                "HORAS", "VALE_REF", "OBSERVACIONES",
                "FECHA_COMPROMISO", "FECHA_PROGRAMADA", "HORAS_ESTIMADAS"]
COLS_OT_ESTADOS = ["ID", "ID_OT", "FECHA_HORA", "ESTADO", "USUARIO", "NOTA"]

# El circuito de una orden. Desde cualquier estado se puede ANULAR.
ESTADOS_OT = ["SOLICITADA", "ASIGNADA", "EN CURSO", "PAUSADA", "RESUELTA", "ANULADA"]
ESTADOS_ABIERTOS = ("SOLICITADA", "ASIGNADA", "EN CURSO", "PAUSADA")
TRANSICIONES_OT = {
    "SOLICITADA": ["ASIGNADA", "ANULADA"],
    "ASIGNADA": ["EN CURSO", "PAUSADA", "RESUELTA", "ANULADA"],
    "EN CURSO": ["PAUSADA", "RESUELTA", "ANULADA"],
    "PAUSADA": ["EN CURSO", "RESUELTA", "ANULADA"],
    "RESUELTA": [],
    "ANULADA": [],
}
PRIORIDADES = ["BAJA", "MEDIA", "ALTA", "URGENTE"]
ORDEN_PRIORIDAD = {"URGENTE": 0, "ALTA": 1, "MEDIA": 2, "BAJA": 3}

# Plazo comprometido según la prioridad, en días corridos desde el alta.
# Se calcula solo al asignar, pero se puede pisar orden por orden.
SLA_DIAS = {"URGENTE": 0, "ALTA": 3, "MEDIA": 7, "BAJA": 15}
HORAS_ESTIMADAS_DEFECTO = 1.0
HORAS_JORNADA = 6  # horas útiles por persona y por día, para medir la carga

# Tipos que puede tener cada renglón de una entrega a un operario.
TIPOS_ENTREGA = ["CONSUMO", "PRESTADO"]
DIAS_PARA_DEMORA = 7

SCOPES = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]

def _admin_inicial():
    """Primer usuario ADMIN, sembrado al crear la hoja Usuarios.

    Sale de los secretos ([admin] en secrets.toml) para no dejar datos
    personales escritos en el código.
    """
    try:
        a = st.secrets["admin"]
        return (a["email"], a.get("nombre", "Administrador"), "ADMIN",
                a.get("sector", "MANTENIMIENTO"))
    except Exception:
        return ("", "Administrador", "ADMIN", "MANTENIMIENTO")


def usando_sheets_reales() -> bool:
    """True si hay credenciales y la app debe trabajar sobre la planilla real.

    Dos frenos para poder probar sin riesgo de tocar los datos de producción:
      - la variable de entorno PANOL_MODO_LOCAL=1
      - la existencia del archivo _devdata/MODO_LOCAL
    Cualquiera de los dos fuerza el modo local aunque haya credenciales.
    """
    if os.environ.get("PANOL_MODO_LOCAL") == "1":
        return False
    if (DEVDATA_DIR / "MODO_LOCAL").exists():
        return False
    try:
        return "gcp_service_account" in st.secrets and "gcp" in st.secrets
    except Exception:
        return False


# ------------------------------------------------------------------ Google Sheets

@st.cache_resource(show_spinner=False)
def _spreadsheet():
    import gspread
    from google.oauth2.service_account import Credentials

    creds = Credentials.from_service_account_info(dict(st.secrets["gcp_service_account"]), scopes=SCOPES)
    return gspread.authorize(creds).open_by_key(st.secrets["gcp"]["sheet_id"])


@st.cache_resource(show_spinner=False)
def _hoja(nombre: str):
    """La hoja, guardada para no volver a pedírsela a Google.

    `ss.worksheet(nombre)` no es gratis: por dentro descarga la ficha completa de
    la planilla. Sin esta caché, cada lectura costaba dos llamadas a la API en
    vez de una, y con eso se llegaba al tope por minuto haciendo unas pocas
    devoluciones seguidas.

    No se limpia junto con el resto de la caché (`limpiar_cache`): lo que cambia
    seguido es el contenido de las hojas, no qué hojas hay.
    """
    return _con_reintentos(lambda: _spreadsheet().worksheet(nombre))


def _ws(nombre: str, crear_con=None):
    """Devuelve la hoja. Si no existe y se pasan encabezados, la crea (solo hojas propias de la app)."""
    from gspread.exceptions import WorksheetNotFound

    try:
        return _hoja(nombre)
    except WorksheetNotFound:
        # Solo se crea la hoja cuando Google confirma que no existe. Antes
        # entraba acá con cualquier error —incluido un corte pasajero— y podía
        # terminar creando una hoja duplicada.
        if crear_con is None:
            raise
        ss = _spreadsheet()
        ws = ss.add_worksheet(title=nombre, rows=1000, cols=max(10, len(crear_con)))
        ws.append_row(crear_con)
        if nombre == HOJA_USUARIOS:
            admin = _admin_inicial()
            if admin[0]:
                ws.append_row([admin[0], admin[1], admin[2], admin[3], "TRUE", ""])
        _hoja.clear()  # ahora sí existe: que la próxima búsqueda la encuentre
        return ws


def a_numero(valor) -> float:
    """Convierte a número un valor que puede venir formateado desde la planilla.

    Google Sheets devuelve el texto tal como se ve: "$ 10.000", "1.000",
    "1.234,56". Hay que distinguir cuándo el punto separa miles y cuándo es
    la coma decimal, o si no "$ 10.000" se leería como diez.
    """
    if valor is None or isinstance(valor, bool):
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)

    t = re.sub(r"[^\d,.\-]", "", str(valor)).strip()
    if not t or t in ("-", ".", ","):
        return 0.0

    tiene_punto, tiene_coma = "." in t, "," in t
    if tiene_punto and tiene_coma:
        # el separador decimal es el que aparece más a la derecha
        if t.rfind(",") > t.rfind("."):
            t = t.replace(".", "").replace(",", ".")
        else:
            t = t.replace(",", "")
    elif tiene_coma:
        t = t.replace(",", ".") if t.count(",") == 1 else t.replace(",", "")
    elif tiene_punto:
        entero, _, decimales = t.rpartition(".")
        # "10.000" son diez mil; "0.125" y "7791.88" son decimales
        miles = (t.count(".") > 1
                 or (len(decimales) == 3 and entero.lstrip("-") not in ("", "0")))
        t = t.replace(".", "") if miles else t

    try:
        return float(t)
    except ValueError:
        return 0.0


def _a_numeros(serie) -> pd.Series:
    return pd.Series([a_numero(v) for v in serie], index=serie.index, dtype="float64")


def _completar_columnas(df: pd.DataFrame, columnas) -> pd.DataFrame:
    """Agrega vacías las columnas esperadas que falten (planillas de versiones previas)."""
    if not columnas:
        return df
    for c in columnas:
        if c not in df.columns:
            df[c] = ""
    return df


def _leer_hoja(nombre: str, crear_con=None, sin_formato=False) -> pd.DataFrame:
    """Lee una hoja completa.

    sin_formato=True pide los valores crudos en vez del texto que se ve. Se usa
    en Inventario, donde los precios se muestran redondeados ("$ 7.792" para
    7791,88) y ese redondeo desviaría la valorización. No conviene en las hojas
    con fechas: ahí los valores crudos vienen como número de serie.
    """
    ws = _ws(nombre, crear_con)
    valores = _con_reintentos(
        lambda: ws.get_all_values(value_render_option="UNFORMATTED_VALUE")
        if sin_formato else ws.get_all_values())
    if len(valores) < 2:
        return pd.DataFrame(columns=valores[0] if valores else (crear_con or []))
    encabezados = valores[0]
    ancho = len(encabezados)
    filas = [(f + [""] * ancho)[:ancho] for f in valores[1:]]
    return _completar_columnas(pd.DataFrame(filas, columns=encabezados), crear_con)


def _col_letra(idx: int) -> str:
    """0 -> A, 25 -> Z, 26 -> AA."""
    letra = ""
    idx += 1
    while idx:
        idx, resto = divmod(idx - 1, 26)
        letra = chr(65 + resto) + letra
    return letra


# ------------------------------------------------------------- Fallback local (dev)

def _leer_local(nombre: str, crear_con=None) -> pd.DataFrame:
    import openpyxl

    csv = DEVDATA_DIR / f"{nombre}.csv"
    if csv.exists():
        return _completar_columnas(pd.read_csv(csv, dtype=str).fillna(""), crear_con)

    if SEED_XLSX.exists():
        wb = openpyxl.load_workbook(SEED_XLSX, data_only=True)
        if nombre in wb.sheetnames:
            ws = wb[nombre]
            valores = [["" if v is None else str(v) for v in r] for r in ws.iter_rows(values_only=True)]
            if nombre == HOJA_PLANO:
                df = pd.DataFrame(valores)
            else:
                encabezados = valores[0]
                ancho = len(encabezados)
                filas = [(f + [""] * ancho)[:ancho] for f in valores[1:]]
                df = pd.DataFrame(filas, columns=encabezados)
                df = df[df.apply(lambda r: any(str(v).strip() for v in r), axis=1)]
                df = _completar_columnas(df, crear_con)
            df.to_csv(csv, index=False)
            return df

    df = pd.DataFrame(columns=crear_con or [])
    if nombre == HOJA_USUARIOS:
        admin = _admin_inicial()
        df = pd.DataFrame([{"EMAIL": admin[0] or "admin@ejemplo.com", "NOMBRE": admin[1],
                            "ROL": admin[2], "SECTOR": admin[3], "ACTIVO": "TRUE",
                            "PASSWORD_HASH": ""}], columns=COLS_USUARIOS)
    DEVDATA_DIR.mkdir(exist_ok=True)
    df.to_csv(csv, index=False)
    return df


def _guardar_local(nombre: str, df: pd.DataFrame):
    DEVDATA_DIR.mkdir(exist_ok=True)
    df.to_csv(DEVDATA_DIR / f"{nombre}.csv", index=False)


def _leer(nombre: str, crear_con=None, sin_formato=False) -> pd.DataFrame:
    if usando_sheets_reales():
        return _leer_hoja(nombre, crear_con, sin_formato)
    return _leer_local(nombre, crear_con)


def _agregar_fila(nombre: str, fila: dict, columnas: list):
    if usando_sheets_reales():
        hoja = _ws(nombre, columnas)
        _con_reintentos(
            lambda: hoja.append_row([str(fila.get(c, "")) for c in columnas],
                                    value_input_option="USER_ENTERED"))
    else:
        df = _leer_local(nombre, columnas)
        nueva = {c: str(fila.get(c, "")) for c in (df.columns if len(df.columns) else columnas)}
        df = pd.concat([df.astype(str), pd.DataFrame([nueva])], ignore_index=True).fillna("")
        _guardar_local(nombre, df)


def limpiar_cache():
    st.cache_data.clear()


# ------------------------------------------------------------------ Inventario

_RANGO_CANT = "'Registro APP'!$D$2:$D$1999"
_RANGO_ITEM = "'Registro APP'!$C$2:$C$1999"
_RANGO_TIPO = "'Registro APP'!$F$2:$F$1999"
_RANGO_DEVUELTA = "'Registro APP'!$K$2:$K$1999"
_RANGO_ESTADO_RENGLON = "'Registro APP'!$L$2:$L$1999"


def _formula_consumo(n: int) -> str:
    """Consumo real de un producto: lo entregado menos el sobrante devuelto."""
    return (f'=SUMIFS({_RANGO_CANT};{_RANGO_ITEM};A{n};{_RANGO_TIPO};"CONSUMO")'
            f'-SUMIFS({_RANGO_DEVUELTA};{_RANGO_ITEM};A{n};{_RANGO_TIPO};"CONSUMO")')


def _formula_prestamo(n: int) -> str:
    """Préstamos sin devolver: entregado menos devuelto, solo renglones pendientes."""
    return (f'=SUMIFS({_RANGO_CANT};{_RANGO_ITEM};A{n};{_RANGO_TIPO};"PRESTADO";'
            f'{_RANGO_ESTADO_RENGLON};"PENDIENTE")'
            f'-SUMIFS({_RANGO_DEVUELTA};{_RANGO_ITEM};A{n};{_RANGO_TIPO};"PRESTADO";'
            f'{_RANGO_ESTADO_RENGLON};"PENDIENTE")')


ESTADOS_STOCK = ["OK", "Mínimo", "Sin stock"]


def _estado(stock_actual: float, stock_minimo: float) -> str:
    """Semáforo del stock. El color lo pone la pantalla, ver estilo.COLORES_STOCK.

    En la planilla la columna Estado/Alerta guarda lo mismo pero con círculos de
    colores, porque ahí no hay forma de pintar la celda según el valor. La app
    calcula el estado por su cuenta y no lee esa columna.
    """
    if stock_actual <= 0:
        return "Sin stock"
    if stock_actual <= stock_minimo:
        return "Mínimo"
    return "OK"


@st.cache_data(ttl=30, show_spinner=False)
def get_items() -> pd.DataFrame:
    # sin formato: los precios se muestran redondeados y eso desviaría el total
    crudo = _leer(HOJA_INVENTARIO, sin_formato=True)
    if crudo.empty:
        return pd.DataFrame(columns=list(COLS_INVENTARIO))

    df = pd.DataFrame()
    for interno, encabezado in COLS_INVENTARIO.items():
        df[interno] = crudo[encabezado] if encabezado in crudo.columns else ""

    df["_fila"] = range(2, len(df) + 2)  # fila real en la planilla (encabezado = 1)
    df = df[df["descripcion"].astype(str).str.strip() != ""]

    for c in ["id", "stock_actual", "stock_minimo", "precio_unitario", "stock_inicial"]:
        df[c] = _a_numeros(df[c])
    df["id"] = df["id"].astype(int)

    for c in ["descripcion", "categoria", "subcategoria", "unidad", "ubicacion",
              "fuente_requerimiento"]:
        df[c] = df[c].astype(str).str.strip()

    df["estado"] = [_estado(a, m) for a, m in zip(df["stock_actual"], df["stock_minimo"])]
    df["valor"] = df["stock_actual"] * df["precio_unitario"]
    return df.reset_index(drop=True)


def diagnostico_catalogo() -> dict:
    """Compara la planilla cruda con lo que la app termina mostrando.

    Existe para explicar de dónde sale una diferencia que confunde seguido: la
    planilla llega al número 492 y el panel dice 486 materiales. No es un error:
    la app cuenta las filas que tienen descripción, no el número más alto. Si en
    algún momento se borraron filas, la numeración queda con huecos.

    Lo que sí es un problema y esto detecta:
      - filas con número pero sin descripción, que la app ignora (materiales
        cargados que no se ven en el sistema)
      - números repetidos, que rompen la edición porque identifican a dos filas
    """
    crudo = _leer(HOJA_INVENTARIO, sin_formato=True)
    vacio = {"filas": 0, "con_descripcion": 0, "sin_descripcion": 0,
             "mayor": 0, "faltantes": [], "repetidos": [], "sin_numero": 0}
    if crudo.empty:
        return vacio

    col_id = COLS_INVENTARIO["id"]
    col_desc = COLS_INVENTARIO["descripcion"]
    if col_id not in crudo.columns or col_desc not in crudo.columns:
        return vacio

    descripcion = crudo[col_desc].astype(str).str.strip()
    numero = _a_numeros(crudo[col_id])

    # una fila cuenta como cargada si tiene número o descripción; el resto son
    # renglones en blanco al final de la hoja
    cargadas = (descripcion != "") | (numero > 0)
    tiene_desc = cargadas & (descripcion != "")

    numeros = [int(n) for n in numero[tiene_desc] if n > 0]
    mayor = max(numeros) if numeros else 0
    vistos, repetidos = set(), set()
    for n in numeros:
        (repetidos if n in vistos else vistos).add(n)

    return {
        "filas": int(cargadas.sum()),
        "con_descripcion": int(tiene_desc.sum()),
        "sin_descripcion": int((cargadas & (descripcion == "")).sum()),
        "sin_numero": int((tiene_desc & (numero <= 0)).sum()),
        "mayor": mayor,
        "faltantes": sorted(set(range(1, mayor + 1)) - vistos),
        "repetidos": sorted(repetidos),
    }


def update_item(item_id: int, **cambios):
    """Actualiza celdas puntuales de un producto, sin pisar las columnas con fórmulas."""
    items = get_items()
    fila = items[items["id"] == item_id]
    if fila.empty:
        raise ValueError(f"No existe el producto {item_id}")
    cambios = {k: v for k, v in cambios.items() if k in INVENTARIO_ESCRIBIBLES}
    if not cambios:
        return

    if usando_sheets_reales():
        ws = _ws(HOJA_INVENTARIO)
        encabezados = ws.row_values(1)
        nro_fila = int(fila.iloc[0]["_fila"])
        actualizaciones = []
        for interno, valor in cambios.items():
            encabezado = COLS_INVENTARIO[interno]
            if encabezado not in encabezados:
                continue
            letra = _col_letra(encabezados.index(encabezado))
            actualizaciones.append({"range": f"{letra}{nro_fila}", "values": [[valor]]})
        if actualizaciones:
            ws.batch_update(actualizaciones, value_input_option="USER_ENTERED")
    else:
        crudo = _leer_local(HOJA_INVENTARIO)
        idx = int(fila.iloc[0]["_fila"]) - 2
        for interno, valor in cambios.items():
            columna = COLS_INVENTARIO[interno]
            crudo[columna] = crudo[columna].astype(str)
            crudo.at[idx, columna] = str(valor)
        _guardar_local(HOJA_INVENTARIO, crudo)

    limpiar_cache()


def add_item(descripcion, categoria, subcategoria, unidad, ubicacion,
             stock_minimo, stock_actual, precio_unitario):
    items = get_items()
    nuevo_id = int(items["id"].max()) + 1 if not items.empty else 1
    n = int(items["_fila"].max()) + 1 if not items.empty else 2  # fila que va a ocupar

    # Las columnas calculadas se cargan como fórmula, igual que el resto de la planilla.
    fila = {
        COLS_INVENTARIO["id"]: nuevo_id,
        COLS_INVENTARIO["descripcion"]: descripcion,
        COLS_INVENTARIO["stock_inicial"]: stock_actual,
        COLS_INVENTARIO["stock_actual"]: f"=C{n}-I{n}-J{n}",
        COLS_INVENTARIO["unidad"]: unidad,
        COLS_INVENTARIO["ubicacion"]: ubicacion,
        COLS_INVENTARIO["estado_hoja"]: f'=IF(D{n}<=0;"🔴 SIN STOCK";IF(D{n}<=H{n};"🟡 MÍNIMO";"🟢 OK"))',
        COLS_INVENTARIO["stock_minimo"]: stock_minimo,
        COLS_INVENTARIO["consumo"]: _formula_consumo(n),
        COLS_INVENTARIO["prestamos_pendiente"]: _formula_prestamo(n),
        COLS_INVENTARIO["precio_unitario"]: precio_unitario,
        COLS_INVENTARIO["total"]: f"=D{n}*K{n}",
        COLS_INVENTARIO["categoria"]: categoria,
        COLS_INVENTARIO["subcategoria"]: subcategoria,
    }
    _agregar_fila(HOJA_INVENTARIO, fila, list(COLS_INVENTARIO.values()))
    limpiar_cache()
    return nuevo_id


# ------------------------------------------------------------------ Vales / movimientos

@st.cache_data(ttl=15, show_spinner=False)
def get_vales() -> pd.DataFrame:
    df = _leer(HOJA_VALES, COLS_VALES)
    if df.empty:
        return pd.DataFrame(columns=COLS_VALES)
    return df[df["ID VALE"].astype(str).str.strip() != ""].reset_index(drop=True)


@st.cache_data(ttl=15, show_spinner=False)
def get_registro() -> pd.DataFrame:
    """Renglones de los vales, con la cantidad pendiente ya calculada."""
    df = _leer(HOJA_REGISTRO, COLS_REGISTRO)
    if df.empty:
        return pd.DataFrame(columns=COLS_REGISTRO + ["pendiente", "_fila"])
    df["_fila"] = range(2, len(df) + 2)  # fila real en la planilla
    df = df[df["ID_REGISTRO"].astype(str).str.strip() != ""].copy()
    for c in ["ID_ITEM", "CANT", "CANT_DEVUELTA"]:
        df[c] = _a_numeros(df[c])
    # Renglones cargados antes de que existiera la columna: un préstamo sin
    # estado se considera pendiente; cualquier otro tipo, ya cerrado.
    df["ESTADO_RENGLON"] = df["ESTADO_RENGLON"].astype(str).str.strip().str.upper()
    sin_estado = df["ESTADO_RENGLON"] == ""
    df.loc[sin_estado, "ESTADO_RENGLON"] = df.loc[sin_estado, "TIPO_MOV"].apply(
        lambda t: "PENDIENTE" if str(t).upper() == "PRESTADO" else "CERRADO")
    df["pendiente"] = (df["CANT"] - df["CANT_DEVUELTA"]).clip(lower=0)
    return df.reset_index(drop=True)


@st.cache_data(ttl=15, show_spinner=False)
def get_movimientos() -> pd.DataFrame:
    """Renglones de Registro APP enriquecidos con los datos de su vale.

    Es la vista que usan el panel, el historial y el historial del operario.
    """
    reg = get_registro()
    if reg.empty:
        return reg

    vales = get_vales()
    if vales.empty:
        for c in ["SECTOR", "ÁREA / SALA", "Receptor / Para Quien", "REGISTRADO_POR"]:
            reg[c] = ""
        return reg

    columnas = ["ID VALE", "SECTOR", "ÁREA / SALA", "Receptor / Para Quien", "REGISTRADO_POR"]
    df = reg.merge(vales[columnas], left_on="ID_VALE_REF", right_on="ID VALE", how="left")
    return df.fillna({"SECTOR": "", "ÁREA / SALA": "", "Receptor / Para Quien": "",
                      "REGISTRADO_POR": ""})


def estado_movimiento(movs: pd.DataFrame) -> pd.Series:
    """Estado de cada renglón en texto, para mostrar en las tablas.

    Un préstamo queda pendiente hasta que vuelve todo. Si volvió una parte, se
    muestra cuánto falta: "Pendiente (3 de 10)".
    """
    if movs.empty:
        return pd.Series(dtype=str)

    def etiqueta(r):
        if str(r["ESTADO_RENGLON"]).strip().upper() != "PENDIENTE":
            return "Cerrado"
        falta, total = float(r["pendiente"]), float(r["CANT"])
        return f"Pendiente ({falta:g} de {total:g})" if falta < total else "Pendiente"

    return movs.apply(etiqueta, axis=1)


FORMATOS_FECHA = ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M",
                  "%d/%m/%Y", "%Y-%m-%d")


def parse_fecha(texto):
    """Interpreta una fecha guardada en la planilla. None si no se puede."""
    t = str(texto).strip()
    if not t:
        return None
    for formato in FORMATOS_FECHA:
        try:
            return dt.datetime.strptime(t, formato)
        except ValueError:
            continue
    return None


def dias_desde(fecha_texto: str) -> int:
    """Días transcurridos desde una fecha de vale. -1 si no se puede interpretar."""
    for formato in ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return (ahora() - dt.datetime.strptime(str(fecha_texto).strip(), formato)).days
        except ValueError:
            continue
    return -1


def _siguiente_id_vale() -> str:
    vales = get_vales()
    nums = pd.to_numeric(
        vales["ID VALE"].astype(str).str.extract(r"(\d+)", expand=False), errors="coerce"
    ).dropna() if not vales.empty else pd.Series(dtype=float)
    return f"V-{int(nums.max()) + 1 if len(nums) else 1:04d}"


def _siguiente_id_registro() -> int:
    reg = get_registro()
    if reg.empty:
        return 1
    nums = pd.to_numeric(reg["ID_REGISTRO"], errors="coerce").dropna()
    return int(nums.max()) + 1 if len(nums) else 1


def registrar_vale(sector, area_sala, receptor, observaciones, renglones, registrado_por=""):
    """Crea un vale (una entrega a un operario) con uno o más renglones.

    Cada renglón lleva su propio tipo: un mismo vale puede tener una herramienta
    prestada y unos tornillos consumidos.

    renglones: lista de dicts con item_id, descripcion, cantidad, unidad, tipo.

    No toca el stock: en la planilla, "Stock Actual" es una fórmula que se
    recalcula sola a partir de estos renglones.
    """
    if not renglones:
        raise ValueError("El vale no tiene renglones")

    id_vale = _siguiente_id_vale()
    momento = ahora_texto()

    tipos = {r["tipo"] for r in renglones}
    tipo_vale = tipos.pop() if len(tipos) == 1 else "MIXTO"
    hay_prestamos = any(r["tipo"] == "PRESTADO" for r in renglones)
    estado_vale = "ABIERTO" if hay_prestamos else "CERRADO"

    _agregar_fila(HOJA_VALES, {
        "ID VALE": id_vale, "FECHA HORA": momento, "TIPO MOVIMIENTO": tipo_vale,
        "SECTOR": sector, "ÁREA / SALA": area_sala, "Receptor / Para Quien": receptor,
        "OBSERVACIONES": observaciones, "ESTADO VALE": estado_vale, "DIAS RETRASO": "",
        "REGISTRADO_POR": registrado_por,
    }, COLS_VALES)

    id_reg = _siguiente_id_registro()
    for r in renglones:
        # un préstamo queda esperando devolución; un consumo nace cerrado
        estado_renglon = "PENDIENTE" if r["tipo"] == "PRESTADO" else "CERRADO"
        _agregar_fila(HOJA_REGISTRO, {
            "ID_REGISTRO": id_reg, "ID_VALE_REF": id_vale, "ID_ITEM": r["item_id"],
            "CANT": r["cantidad"], "UNIDAD": r.get("unidad", ""), "TIPO_MOV": r["tipo"],
            "DESCRIPCIÓN_ITEM": r["descripcion"], "FECHA_VALE": momento,
            "OBSERVACIONES": observaciones, "ESTADO_VALE (auto)": estado_vale,
            "CANT_DEVUELTA": 0, "ESTADO_RENGLON": estado_renglon,
        }, COLS_REGISTRO)
        id_reg += 1

    limpiar_cache()
    return id_vale


def _escribir_celdas(hoja: str, columnas: list, nro_fila: int, cambios: dict):
    """Escribe celdas puntuales de una fila, sin pisar el resto."""
    cambios = {k: v for k, v in cambios.items() if k in columnas}
    if not cambios:
        return
    if usando_sheets_reales():
        ws = _ws(hoja, columnas)
        _con_reintentos(lambda: ws.batch_update(
            [{"range": f"{_col_letra(columnas.index(c))}{nro_fila}", "values": [[v]]}
             for c, v in cambios.items()],
            value_input_option="USER_ENTERED"))
    else:
        df = _leer_local(hoja, columnas)
        idx = nro_fila - 2
        for c, v in cambios.items():
            df[c] = df[c].astype(str)
            df.at[idx, c] = str(v)
        _guardar_local(hoja, df)


def _escribir_celda_registro(nro_fila: int, columna: str, valor):
    """Escribe una celda puntual de un renglón de Registro APP."""
    _escribir_celdas(HOJA_REGISTRO, COLS_REGISTRO, nro_fila, {columna: valor})


def _recalcular_estado_vale(id_vale: str):
    """Cierra el vale cuando ya no le queda ningún renglón pendiente."""
    reg = get_registro()
    renglones = reg[reg["ID_VALE_REF"] == id_vale]
    if renglones.empty:
        return
    sigue_abierto = (renglones["ESTADO_RENGLON"] == "PENDIENTE").any()
    estado = "ABIERTO" if sigue_abierto else "CERRADO"

    if usando_sheets_reales():
        ws = _ws(HOJA_VALES)
        celda = ws.find(id_vale, in_column=1)
        if celda:
            letra = _col_letra(COLS_VALES.index("ESTADO VALE"))
            ws.update(values=[[estado]], range_name=f"{letra}{celda.row}")
    else:
        vales = _leer_local(HOJA_VALES, COLS_VALES)
        vales.loc[vales["ID VALE"] == id_vale, "ESTADO VALE"] = estado
        _guardar_local(HOJA_VALES, vales)

    for _, r in renglones.iterrows():
        _escribir_celda_registro(int(r["_fila"]), "ESTADO_VALE (auto)", estado)
    limpiar_cache()


def devolver_renglon(id_registro, cantidad: float):
    """Registra la devolución (total o parcial) de un renglón.

    Sirve tanto para una herramienta prestada como para el sobrante de un
    consumo. El stock se repone solo, por la fórmula de la planilla.
    """
    reg = get_registro()
    fila = reg[reg["ID_REGISTRO"].astype(str) == str(id_registro)]
    if fila.empty:
        raise ValueError(f"No existe el renglón {id_registro}")
    fila = fila.iloc[0]

    pendiente = float(fila["pendiente"])
    cantidad = float(cantidad)
    if cantidad <= 0:
        raise ValueError("La cantidad a devolver tiene que ser mayor que cero")
    if cantidad > pendiente:
        raise ValueError(f"No podés devolver más de lo pendiente ({pendiente:g})")

    nueva_devuelta = float(fila["CANT_DEVUELTA"]) + cantidad
    nro_fila = int(fila["_fila"])
    _escribir_celda_registro(nro_fila, "CANT_DEVUELTA", nueva_devuelta)
    if nueva_devuelta >= float(fila["CANT"]):
        _escribir_celda_registro(nro_fila, "ESTADO_RENGLON", "CERRADO")

    limpiar_cache()
    _recalcular_estado_vale(fila["ID_VALE_REF"])


def convertir_a_consumo(id_registro):
    """El préstamo no vuelve (se perdió, se rompió o se lo quedaron).

    Pasa el renglón a CONSUMO y lo cierra. No mueve el stock: préstamo y
    consumo descuentan igual, solo cambia si se espera la devolución.
    """
    reg = get_registro()
    fila = reg[reg["ID_REGISTRO"].astype(str) == str(id_registro)]
    if fila.empty:
        raise ValueError(f"No existe el renglón {id_registro}")
    fila = fila.iloc[0]

    nro_fila = int(fila["_fila"])
    _escribir_celda_registro(nro_fila, "TIPO_MOV", "CONSUMO")
    _escribir_celda_registro(nro_fila, "ESTADO_RENGLON", "CERRADO")

    limpiar_cache()
    _recalcular_estado_vale(fila["ID_VALE_REF"])


def cerrar_vale(id_vale: str):
    """Devuelve de una todo lo que quede pendiente en el vale."""
    reg = get_registro()
    pendientes = reg[(reg["ID_VALE_REF"] == id_vale) & (reg["ESTADO_RENGLON"] == "PENDIENTE")]
    for _, r in pendientes.iterrows():
        if r["pendiente"] > 0:
            devolver_renglon(r["ID_REGISTRO"], r["pendiente"])
    _recalcular_estado_vale(id_vale)


def registrar_ingreso(item_id, cantidad, observaciones, responsable):
    """Reposición del pañol: suma al Stock Inicial y deja constancia del ingreso.

    Provisional hasta que rediseñemos la sección de Ingresos: las fórmulas de la
    planilla solo restan consumos y préstamos, no suman ingresos, así que la
    entrada se carga sobre "Stock Inicial" y el renglón queda como historial.
    """
    items = get_items()
    fila = items[items["id"] == int(item_id)]
    if fila.empty:
        raise ValueError(f"No existe el producto {item_id}")
    fila = fila.iloc[0]

    id_vale = _siguiente_id_vale()
    momento = ahora_texto()
    _agregar_fila(HOJA_VALES, {
        "ID VALE": id_vale, "FECHA HORA": momento, "TIPO MOVIMIENTO": "INGRESO",
        "SECTOR": "", "ÁREA / SALA": "", "Receptor / Para Quien": responsable,
        "OBSERVACIONES": observaciones, "ESTADO VALE": "CERRADO", "DIAS RETRASO": "",
        "REGISTRADO_POR": responsable,
    }, COLS_VALES)
    _agregar_fila(HOJA_REGISTRO, {
        "ID_REGISTRO": _siguiente_id_registro(), "ID_VALE_REF": id_vale,
        "ID_ITEM": int(item_id), "CANT": cantidad, "UNIDAD": fila["unidad"],
        "TIPO_MOV": "INGRESO", "DESCRIPCIÓN_ITEM": fila["descripcion"],
        "FECHA_VALE": momento, "OBSERVACIONES": observaciones,
        "ESTADO_VALE (auto)": "CERRADO", "CANT_DEVUELTA": 0, "ESTADO_RENGLON": "CERRADO",
    }, COLS_REGISTRO)

    update_item(int(item_id), stock_inicial=float(fila["stock_inicial"]) + float(cantidad))
    limpiar_cache()
    return id_vale


# ------------------------------------------------------------------ Usuarios

@st.cache_data(ttl=30, show_spinner=False)
def get_usuarios() -> pd.DataFrame:
    df = _leer(HOJA_USUARIOS, COLS_USUARIOS)
    if df.empty:
        return pd.DataFrame(columns=COLS_USUARIOS)
    for c in COLS_USUARIOS:
        if c not in df.columns:
            df[c] = ""
    df = df[df["EMAIL"].astype(str).str.strip() != ""]
    return df[COLS_USUARIOS].reset_index(drop=True)


def get_usuarios_activos() -> list:
    df = get_usuarios()
    if df.empty:
        return []
    activos = df[df["ACTIVO"].astype(str).str.upper().isin(["TRUE", "1", "SI", "SÍ", "VERDADERO"])]
    return activos.to_dict("records")


def add_usuario(email, nombre, rol, sector, password_hash=""):
    _agregar_fila(HOJA_USUARIOS, {
        "EMAIL": email.strip().lower(), "NOMBRE": nombre, "ROL": rol,
        "SECTOR": sector, "ACTIVO": "TRUE", "PASSWORD_HASH": password_hash,
    }, COLS_USUARIOS)
    limpiar_cache()


def set_password_hash(email: str, password_hash: str):
    email = email.strip().lower()
    if usando_sheets_reales():
        ws = _ws(HOJA_USUARIOS, COLS_USUARIOS)
        encabezados = ws.row_values(1)
        celda = ws.find(email, in_column=1)
        if celda is None:
            raise ValueError(f"No se encontró el usuario {email}")
        letra = _col_letra(encabezados.index("PASSWORD_HASH"))
        ws.update(f"{letra}{celda.row}", [[password_hash]])
    else:
        df = _leer_local(HOJA_USUARIOS, COLS_USUARIOS)
        df.loc[df["EMAIL"].astype(str).str.lower() == email, "PASSWORD_HASH"] = password_hash
        _guardar_local(HOJA_USUARIOS, df)
    limpiar_cache()


def set_usuario_activo(email: str, activo: bool):
    email = email.strip().lower()
    valor = "TRUE" if activo else "FALSE"
    if usando_sheets_reales():
        ws = _ws(HOJA_USUARIOS, COLS_USUARIOS)
        encabezados = ws.row_values(1)
        celda = ws.find(email, in_column=1)
        if celda:
            ws.update(f"{_col_letra(encabezados.index('ACTIVO'))}{celda.row}", [[valor]])
    else:
        df = _leer_local(HOJA_USUARIOS, COLS_USUARIOS)
        df.loc[df["EMAIL"].astype(str).str.lower() == email, "ACTIVO"] = valor
        _guardar_local(HOJA_USUARIOS, df)
    limpiar_cache()


# ------------------------------------------------------------------ Reclamos

@st.cache_data(ttl=15, show_spinner=False)
def get_reclamos() -> pd.DataFrame:
    df = _leer(HOJA_RECLAMOS, COLS_RECLAMOS)
    if df.empty:
        return pd.DataFrame(columns=COLS_RECLAMOS)
    return df[df["ID"].astype(str).str.strip() != ""].reset_index(drop=True)


def add_reclamo(tipo, producto, detalle, email, nombre):
    reclamos = get_reclamos()
    nums = pd.to_numeric(reclamos["ID"], errors="coerce").dropna() if not reclamos.empty else pd.Series(dtype=float)
    nuevo_id = int(nums.max()) + 1 if len(nums) else 1
    _agregar_fila(HOJA_RECLAMOS, {
        "ID": nuevo_id, "FECHA_HORA": ahora_texto(),
        "TIPO": tipo, "PRODUCTO": producto, "DETALLE": detalle,
        "EMAIL": email, "NOMBRE": nombre, "ESTADO": "ABIERTO", "RESPUESTA": "",
    }, COLS_RECLAMOS)
    limpiar_cache()


def responder_reclamo(reclamo_id, estado, respuesta):
    if usando_sheets_reales():
        ws = _ws(HOJA_RECLAMOS, COLS_RECLAMOS)
        encabezados = ws.row_values(1)
        celda = ws.find(str(reclamo_id), in_column=1)
        if celda:
            ws.batch_update([
                {"range": f"{_col_letra(encabezados.index('ESTADO'))}{celda.row}", "values": [[estado]]},
                {"range": f"{_col_letra(encabezados.index('RESPUESTA'))}{celda.row}", "values": [[respuesta]]},
            ], value_input_option="USER_ENTERED")
    else:
        df = _leer_local(HOJA_RECLAMOS, COLS_RECLAMOS)
        mask = df["ID"].astype(str) == str(reclamo_id)
        df.loc[mask, "ESTADO"] = estado
        df.loc[mask, "RESPUESTA"] = respuesta
        _guardar_local(HOJA_RECLAMOS, df)
    limpiar_cache()


# ------------------------------------------------------------------ Parámetros y plano

@st.cache_data(ttl=300, show_spinner=False)
def get_parametros() -> dict:
    """Devuelve {columna: [valores no vacíos]} de la hoja Parametros."""
    df = _leer(HOJA_PARAMETROS)
    if df.empty:
        return {}
    return {c: [v for v in df[c].astype(str).str.strip().tolist() if v] for c in df.columns}


@st.cache_data(ttl=300, show_spinner=False)
def get_estanterias() -> pd.DataFrame:
    """Catálogo de estanterías del pañol (número, medidas, área y qué guarda)."""
    crudo = _leer(HOJA_PLANO)
    if crudo.empty:
        return pd.DataFrame(columns=["estanteria", "ancho", "profundidad", "estantes", "area", "objetos"])

    filas = []
    for fila in crudo.itertuples(index=False):
        vals = ["" if pd.isna(v) else str(v).strip() for v in fila]
        vals += [""] * (10 - len(vals))
        estanteria = vals[1]
        if not estanteria or estanteria.lower().startswith("estanteria"):
            continue
        filas.append({
            "estanteria": estanteria.zfill(2) if estanteria.isdigit() else estanteria,
            "ancho": vals[2], "profundidad": vals[3], "estantes": vals[4],
            "area": vals[5], "objetos": vals[9],
        })
    return pd.DataFrame(filas)


def numero_estanteria(ubicacion: str) -> str:
    """Extrae el número de estantería de una ubicación libre ('18', '18-2', 'Est. 18')."""
    import re

    m = re.search(r"\d+", str(ubicacion or ""))
    return m.group(0).zfill(2) if m else ""


# ══════════════════════════════════ Órdenes de trabajo (correctivo) ══════════

@st.cache_data(ttl=15, show_spinner=False)
def get_ordenes() -> pd.DataFrame:
    """Órdenes de trabajo, con la fila real de la planilla para poder editarlas."""
    df = _leer(HOJA_ORDENES, COLS_ORDENES)
    if df.empty:
        return pd.DataFrame(columns=COLS_ORDENES + ["_fila", "dias_abierta"])
    df["_fila"] = range(2, len(df) + 2)
    df = df[df["ID_OT"].astype(str).str.strip() != ""].copy()
    if df.empty:
        return pd.DataFrame(columns=COLS_ORDENES + ["_fila", "dias_abierta"])
    df["ESTADO"] = df["ESTADO"].astype(str).str.strip().str.upper().replace("", "SOLICITADA")
    df["PRIORIDAD"] = df["PRIORIDAD"].astype(str).str.strip().str.upper()
    df["HORAS"] = _a_numeros(df["HORAS"])
    df["HORAS_ESTIMADAS"] = _a_numeros(df["HORAS_ESTIMADAS"])
    df["dias_abierta"] = df["FECHA_ALTA"].apply(dias_desde)
    df["orden_prioridad"] = df["PRIORIDAD"].map(ORDEN_PRIORIDAD).fillna(9)

    dia = hoy()
    # se arman como listas y no con .apply(): pandas convertiría los None en NaT
    # y después NaT no se puede restar con una fecha
    compromisos = [parse_fecha(v) for v in df["FECHA_COMPROMISO"]]
    df["dias_para_vencer"] = [
        (f.date() - dia).days if f is not None else None for f in compromisos]
    abierta = df["ESTADO"].isin(ESTADOS_ABIERTOS)
    df["vencida"] = [
        bool(a and d is not None and d < 0)
        for a, d in zip(abierta, df["dias_para_vencer"])]
    df["vence_hoy"] = [
        bool(a and d == 0) for a, d in zip(abierta, df["dias_para_vencer"])]

    df["dia_programado"] = [
        f.date() if f is not None else None
        for f in (parse_fecha(v) for v in df["FECHA_PROGRAMADA"])]
    return df.reset_index(drop=True)


def calcular_compromiso(prioridad: str, desde=None) -> str:
    """Fecha comprometida según la prioridad. Devuelve texto listo para la planilla."""
    base = desde or ahora()
    dias = SLA_DIAS.get(str(prioridad).strip().upper(), 7)
    return (base + dt.timedelta(days=dias)).strftime("%Y-%m-%d")


@st.cache_data(ttl=15, show_spinner=False)
def get_ot_estados() -> pd.DataFrame:
    """Bitácora de cambios de estado de las órdenes."""
    df = _leer(HOJA_OT_ESTADOS, COLS_OT_ESTADOS)
    if df.empty:
        return pd.DataFrame(columns=COLS_OT_ESTADOS)
    return df[df["ID_OT"].astype(str).str.strip() != ""].reset_index(drop=True)


def _siguiente_id_ot() -> str:
    ordenes = get_ordenes()
    if ordenes.empty:
        return "OT-0001"
    nums = pd.to_numeric(
        ordenes["ID_OT"].astype(str).str.extract(r"(\d+)", expand=False), errors="coerce").dropna()
    return f"OT-{int(nums.max()) + 1 if len(nums) else 1:04d}"


def _anotar_estado(id_ot: str, estado: str, usuario: str, nota: str = ""):
    estados = get_ot_estados()
    nums = pd.to_numeric(estados["ID"], errors="coerce").dropna() if not estados.empty else pd.Series(dtype=float)
    _agregar_fila(HOJA_OT_ESTADOS, {
        "ID": int(nums.max()) + 1 if len(nums) else 1,
        "ID_OT": id_ot,
        "FECHA_HORA": ahora_texto(),
        "ESTADO": estado, "USUARIO": usuario, "NOTA": nota,
    }, COLS_OT_ESTADOS)


def crear_solicitud(area, descripcion, prioridad, solicitante, solicitante_email,
                    observaciones=""):
    """Alta de una solicitud de reparación. Nace como orden en estado SOLICITADA."""
    id_ot = _siguiente_id_ot()
    momento = ahora_texto()
    _agregar_fila(HOJA_ORDENES, {
        "ID_OT": id_ot, "FECHA_ALTA": momento, "SOLICITANTE": solicitante,
        "SOLICITANTE_EMAIL": solicitante_email, "AREA": area, "DESCRIPCION": descripcion,
        "PRIORIDAD": prioridad, "SECTOR_ASIGNADO": "", "ASIGNADO_A": "",
        "ESTADO": "SOLICITADA", "FECHA_ASIGNACION": "", "FECHA_CIERRE": "",
        "TRABAJO_REALIZADO": "", "CAUSA": "", "HORAS": "", "VALE_REF": "",
        "OBSERVACIONES": observaciones,
    }, COLS_ORDENES)
    _anotar_estado(id_ot, "SOLICITADA", solicitante, descripcion[:120])
    limpiar_cache()
    return id_ot


def _fila_orden(id_ot: str):
    ordenes = get_ordenes()
    fila = ordenes[ordenes["ID_OT"].astype(str) == str(id_ot)]
    if fila.empty:
        raise ValueError(f"No existe la orden {id_ot}")
    return fila.iloc[0]


def asignar_orden(id_ot, sector, asignado_a, prioridad, usuario, nota="",
                  fecha_compromiso=None, fecha_programada=None, horas_estimadas=None):
    """Asigna la orden a un sector y a una persona, y la pasa a ASIGNADA.

    Si no se indica fecha de compromiso, se calcula sola a partir de la
    prioridad (ver SLA_DIAS). Se puede pisar pasando fecha_compromiso.
    """
    fila = _fila_orden(id_ot)
    cambios = {"SECTOR_ASIGNADO": sector, "ASIGNADO_A": asignado_a, "PRIORIDAD": prioridad}

    if fecha_compromiso:
        cambios["FECHA_COMPROMISO"] = fecha_compromiso
    elif not str(fila["FECHA_COMPROMISO"]).strip() or fila["PRIORIDAD"] != prioridad:
        # sin plazo todavía, o cambió la prioridad: se recalcula
        cambios["FECHA_COMPROMISO"] = calcular_compromiso(
            prioridad, parse_fecha(fila["FECHA_ALTA"]))
    if fecha_programada is not None:
        cambios["FECHA_PROGRAMADA"] = fecha_programada
    if horas_estimadas is not None:
        cambios["HORAS_ESTIMADAS"] = horas_estimadas
    elif not float(fila["HORAS_ESTIMADAS"] or 0):
        cambios["HORAS_ESTIMADAS"] = HORAS_ESTIMADAS_DEFECTO

    if fila["ESTADO"] == "SOLICITADA":
        cambios["ESTADO"] = "ASIGNADA"
        cambios["FECHA_ASIGNACION"] = ahora_texto()

    _escribir_celdas(HOJA_ORDENES, COLS_ORDENES, int(fila["_fila"]), cambios)
    _anotar_estado(id_ot, cambios.get("ESTADO", fila["ESTADO"]), usuario,
                   nota or f"Asignada a {asignado_a} ({sector})")
    limpiar_cache()


def programar_orden(id_ot, fecha_programada, usuario, horas_estimadas=None):
    """Agenda la orden para un día. Pasar '' desaparece de la agenda."""
    fila = _fila_orden(id_ot)
    cambios = {"FECHA_PROGRAMADA": fecha_programada or ""}
    if horas_estimadas is not None:
        cambios["HORAS_ESTIMADAS"] = horas_estimadas
    _escribir_celdas(HOJA_ORDENES, COLS_ORDENES, int(fila["_fila"]), cambios)
    limpiar_cache()


def cambiar_estado_orden(id_ot, nuevo_estado, usuario, nota=""):
    """Mueve la orden a otro estado, validando que la transición sea posible."""
    fila = _fila_orden(id_ot)
    actual = fila["ESTADO"]
    if nuevo_estado not in TRANSICIONES_OT.get(actual, []):
        raise ValueError(f"No se puede pasar de {actual} a {nuevo_estado}")

    cambios = {"ESTADO": nuevo_estado}
    if nuevo_estado in ("RESUELTA", "ANULADA"):
        cambios["FECHA_CIERRE"] = ahora_texto()
    _escribir_celdas(HOJA_ORDENES, COLS_ORDENES, int(fila["_fila"]), cambios)
    _anotar_estado(id_ot, nuevo_estado, usuario, nota)
    limpiar_cache()


def cerrar_orden(id_ot, trabajo_realizado, causa, horas, usuario, nota=""):
    """Cierre técnico: qué se hizo, por qué pasó y cuánto llevó."""
    fila = _fila_orden(id_ot)
    if fila["ESTADO"] not in ("ASIGNADA", "EN CURSO", "PAUSADA"):
        raise ValueError(f"Una orden {fila['ESTADO']} no se puede cerrar")
    if not str(trabajo_realizado).strip():
        raise ValueError("Contá qué trabajo se hizo antes de cerrar")

    _escribir_celdas(HOJA_ORDENES, COLS_ORDENES, int(fila["_fila"]), {
        "ESTADO": "RESUELTA",
        "FECHA_CIERRE": ahora_texto(),
        "TRABAJO_REALIZADO": trabajo_realizado, "CAUSA": causa, "HORAS": horas,
    })
    _anotar_estado(id_ot, "RESUELTA", usuario, nota or trabajo_realizado[:120])
    limpiar_cache()


def indicadores_mantenimiento(ordenes: pd.DataFrame) -> dict:
    """Métricas de gestión para el tablero de jefatura."""
    if ordenes.empty:
        return {"total": 0, "abiertas": 0, "vencidas": 0, "vence_hoy": 0,
                "sin_asignar": 0, "resueltas": 0, "dias_promedio": None,
                "cumplimiento": None, "horas_cerradas": 0}

    abiertas = ordenes[ordenes["ESTADO"].isin(ESTADOS_ABIERTOS)]
    resueltas = ordenes[ordenes["ESTADO"] == "RESUELTA"]

    # cuánto tardó cada orden resuelta, de punta a punta
    duraciones, en_plazo = [], []
    for _, o in resueltas.iterrows():
        alta, cierre = parse_fecha(o["FECHA_ALTA"]), parse_fecha(o["FECHA_CIERRE"])
        if alta and cierre:
            duraciones.append((cierre - alta).total_seconds() / 86400)
        compromiso = parse_fecha(o["FECHA_COMPROMISO"])
        if cierre and compromiso:
            en_plazo.append(cierre.date() <= compromiso.date())

    return {
        "total": len(ordenes),
        "abiertas": len(abiertas),
        "vencidas": int(ordenes["vencida"].sum()),
        "vence_hoy": int(ordenes["vence_hoy"].sum()),
        "sin_asignar": int((ordenes["ESTADO"] == "SOLICITADA").sum()),
        "resueltas": len(resueltas),
        "dias_promedio": round(sum(duraciones) / len(duraciones), 1) if duraciones else None,
        "cumplimiento": round(100 * sum(en_plazo) / len(en_plazo)) if en_plazo else None,
        "horas_cerradas": float(resueltas["HORAS"].sum()),
    }


def carga_por_persona(ordenes: pd.DataFrame) -> pd.DataFrame:
    """Horas pendientes por responsable, para ver quién está saturado."""
    if ordenes.empty:
        return pd.DataFrame(columns=["persona", "ordenes", "horas", "vencidas", "capacidad"])

    abiertas = ordenes[ordenes["ESTADO"].isin(ESTADOS_ABIERTOS) &
                       (ordenes["ASIGNADO_A"].astype(str).str.strip() != "")]
    if abiertas.empty:
        return pd.DataFrame(columns=["persona", "ordenes", "horas", "vencidas", "capacidad"])

    resumen = abiertas.groupby("ASIGNADO_A").agg(
        ordenes=("ID_OT", "count"),
        horas=("HORAS_ESTIMADAS", "sum"),
        vencidas=("vencida", "sum"),
    ).reset_index().rename(columns={"ASIGNADO_A": "persona"})
    # capacidad = qué parte de una semana laboral ocupan esas horas
    resumen["capacidad"] = (resumen["horas"] / (HORAS_JORNADA * 5)).round(2)
    return resumen.sort_values("horas", ascending=False)
